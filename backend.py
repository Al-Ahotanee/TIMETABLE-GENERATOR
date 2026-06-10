# backend.py – Complete: Auth, Academic Hierarchy, AI Timetable Engine, Exports
import os, datetime, logging, re, json, io, csv
from functools import wraps
from flask import Flask, request, jsonify, g, Blueprint, send_file
from flask_cors import CORS
from flask_jwt_extended import (
    JWTManager, create_access_token, create_refresh_token,
    jwt_required, get_jwt_identity, get_jwt
)
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from itsdangerous import URLSafeTimedSerializer
import secrets

# AI & Export libraries
from ortools.sat.python import cp_model
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import PatternFill, Font
from PIL import Image, ImageDraw, ImageFont

# ---------- Initialise extensions ----------
db = SQLAlchemy()
jwt = JWTManager()

# ---------- App Factory ----------
def create_app():
    app = Flask(__name__)
    app.config.update(
        SECRET_KEY=os.environ.get('SECRET_KEY', secrets.token_hex(32)),
        SQLALCHEMY_DATABASE_URI=os.environ.get('DATABASE_URL', 'postgresql://localhost/timetable'),
        SQLALCHEMY_TRACK_MODIFICATIONS=False,
        JWT_SECRET_KEY=os.environ.get('JWT_SECRET_KEY', secrets.token_hex(32)),
        JWT_ACCESS_TOKEN_EXPIRES=datetime.timedelta(hours=1),
        JWT_REFRESH_TOKEN_EXPIRES=datetime.timedelta(days=30),
        MAIL_SERVER=os.environ.get('MAIL_SERVER', 'smtp.gmail.com'),
        MAIL_PORT=int(os.environ.get('MAIL_PORT', 587)),
        MAIL_USE_TLS=os.environ.get('MAIL_USE_TLS', 'true').lower() == 'true',
        MAIL_USERNAME=os.environ.get('MAIL_USERNAME', ''),
        MAIL_PASSWORD=os.environ.get('MAIL_PASSWORD', ''),
        MAIL_DEFAULT_SENDER=os.environ.get('MAIL_DEFAULT_SENDER', 'noreply@timetable.com'),
        TOKEN_SALT=os.environ.get('TOKEN_SALT', 'email-verify'),
    )
    CORS(app, resources={r"/api/*": {"origins": "*"}})
    db.init_app(app)
    jwt.init_app(app)

    # Register all blueprints
    app.register_blueprint(auth_bp, url_prefix='/api/auth')
    app.register_blueprint(admin_bp, url_prefix='/api/admin')
    app.register_blueprint(institution_bp, url_prefix='/api/institution')
    app.register_blueprint(timetable_bp, url_prefix='/api/timetable')

    with app.app_context():
        db.create_all()
        seed_super_admin()

    return app

# ---------- Database Models ----------
class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(256), nullable=False)
    first_name = db.Column(db.String(100))
    last_name = db.Column(db.String(100))
    role = db.Column(db.String(30), nullable=False, default='student')
    institution_id = db.Column(db.Integer, db.ForeignKey('institutions.id'), nullable=True)
    is_verified = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)
    institution = db.relationship('Institution', backref='users')
    lecturer_profile = db.relationship('Lecturer', backref='user', uselist=False)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def to_dict(self):
        return {
            'id': self.id, 'email': self.email, 'first_name': self.first_name,
            'last_name': self.last_name, 'role': self.role,
            'institution_id': self.institution_id, 'is_verified': self.is_verified
        }

class Institution(db.Model):
    __tablename__ = 'institutions'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    code = db.Column(db.String(20), unique=True)
    type = db.Column(db.String(30))
    address = db.Column(db.Text)
    logo_url = db.Column(db.String(500))
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    faculties = db.relationship('Faculty', backref='institution', lazy='dynamic')
    departments = db.relationship('Department', backref='institution', lazy='dynamic')

class AuditLog(db.Model):
    __tablename__ = 'audit_logs'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    action = db.Column(db.String(50), nullable=False)
    details = db.Column(db.Text)
    ip_address = db.Column(db.String(45))
    timestamp = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    user = db.relationship('User', backref='audit_logs')

class Faculty(db.Model):
    __tablename__ = 'faculties'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    code = db.Column(db.String(20))
    institution_id = db.Column(db.Integer, db.ForeignKey('institutions.id'), nullable=False)
    departments = db.relationship('Department', backref='faculty', lazy='dynamic')

class Department(db.Model):
    __tablename__ = 'departments'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    code = db.Column(db.String(20))
    institution_id = db.Column(db.Integer, db.ForeignKey('institutions.id'), nullable=False)
    faculty_id = db.Column(db.Integer, db.ForeignKey('faculties.id'), nullable=False)

class Programme(db.Model):
    __tablename__ = 'programmes'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    code = db.Column(db.String(20))
    department_id = db.Column(db.Integer, db.ForeignKey('departments.id'), nullable=False)
    duration_years = db.Column(db.Integer, default=4)

class AcademicSession(db.Model):
    __tablename__ = 'academic_sessions'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    institution_id = db.Column(db.Integer, db.ForeignKey('institutions.id'), nullable=False)
    start_date = db.Column(db.Date)
    end_date = db.Column(db.Date)
    is_current = db.Column(db.Boolean, default=False)

class Semester(db.Model):
    __tablename__ = 'semesters'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    session_id = db.Column(db.Integer, db.ForeignKey('academic_sessions.id'), nullable=False)
    start_date = db.Column(db.Date)
    end_date = db.Column(db.Date)

class Course(db.Model):
    __tablename__ = 'courses'
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), nullable=False)
    title = db.Column(db.String(200), nullable=False)
    credit_units = db.Column(db.Integer, default=3)
    level = db.Column(db.Integer)
    department_id = db.Column(db.Integer, db.ForeignKey('departments.id'), nullable=False)
    programme_id = db.Column(db.Integer, db.ForeignKey('programmes.id'), nullable=False)
    semester_id = db.Column(db.Integer, db.ForeignKey('semesters.id'), nullable=False)
    lecturer_id = db.Column(db.Integer, db.ForeignKey('lecturers.id'), nullable=True)
    students_count = db.Column(db.Integer, default=30)

class Lecturer(db.Model):
    __tablename__ = 'lecturers'
    id = db.Column(db.Integer, primary_key=True)
    staff_id = db.Column(db.String(50), unique=True, nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), unique=True, nullable=False)
    department_id = db.Column(db.Integer, db.ForeignKey('departments.id'), nullable=False)
    rank = db.Column(db.String(100))
    max_daily_hours = db.Column(db.Integer, default=6)
    max_weekly_hours = db.Column(db.Integer, default=24)
    availability_json = db.Column(db.Text, default='{}')

class Venue(db.Model):
    __tablename__ = 'venues'
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(50), unique=True, nullable=False)
    name = db.Column(db.String(200))
    capacity = db.Column(db.Integer, nullable=False)
    building = db.Column(db.String(100))
    resources = db.Column(db.Text)
    institution_id = db.Column(db.Integer, db.ForeignKey('institutions.id'), nullable=False)

class TimetableGeneration(db.Model):
    __tablename__ = 'timetable_generations'
    id = db.Column(db.Integer, primary_key=True)
    session_id = db.Column(db.Integer, db.ForeignKey('academic_sessions.id'), nullable=False)
    semester_id = db.Column(db.Integer, db.ForeignKey('semesters.id'), nullable=False)
    status = db.Column(db.String(20), default='draft')
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    quality_score = db.Column(db.Float, default=0.0)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

class TimetableSlot(db.Model):
    __tablename__ = 'timetable_slots'
    id = db.Column(db.Integer, primary_key=True)
    generation_id = db.Column(db.Integer, db.ForeignKey('timetable_generations.id'), nullable=True)
    day = db.Column(db.String(10), nullable=False)
    period = db.Column(db.Integer, nullable=False)
    start_time = db.Column(db.Time)
    end_time = db.Column(db.Time)
    course_id = db.Column(db.Integer, db.ForeignKey('courses.id'), nullable=False)
    lecturer_id = db.Column(db.Integer, db.ForeignKey('lecturers.id'), nullable=False)
    venue_id = db.Column(db.Integer, db.ForeignKey('venues.id'), nullable=False)
    session_id = db.Column(db.Integer, db.ForeignKey('academic_sessions.id'), nullable=False)
    semester_id = db.Column(db.Integer, db.ForeignKey('semesters.id'), nullable=False)
    department_id = db.Column(db.Integer, db.ForeignKey('departments.id'), nullable=True)

class Notification(db.Model):
    __tablename__ = 'notifications'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    message = db.Column(db.Text, nullable=False)
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

# ---------- Helper functions ----------
def log_audit(action, user_id=None, details=None):
    try:
        ip = request.remote_addr if request else None
        db.session.add(AuditLog(user_id=user_id, action=action, details=details, ip_address=ip))
        db.session.commit()
    except Exception as e:
        logging.error(f"Audit log failed: {e}")

def role_required(*roles):
    def wrapper(fn):
        @wraps(fn)
        @jwt_required()
        def decorated(*args, **kwargs):
            claims = get_jwt()
            user_role = claims.get('role')
            if user_role not in roles:
                return jsonify({"msg": "Insufficient privileges"}), 403
            g.current_user_id = get_jwt_identity()
            g.current_user = User.query.get(g.current_user_id)
            return fn(*args, **kwargs)
        return decorated
    return wrapper

def get_institution_id():
    user = g.current_user
    if user.role == 'super_admin':
        return None
    return user.institution_id

def inst_query(model):
    inst_id = get_institution_id()
    if inst_id is not None:
        return model.query.filter_by(institution_id=inst_id)
    return model.query

# ---------- Auth Blueprint ----------
auth_bp = Blueprint('auth', __name__)

@auth_bp.route('/register', methods=['POST'])
def register():
    data = request.get_json()
    if not data or not data.get('email') or not data.get('password'):
        return jsonify({"msg": "Email and password required"}), 400
    if not re.match(r"[^@]+@[^@]+\.[^@]+", data['email']):
        return jsonify({"msg": "Invalid email format"}), 400
    if User.query.filter_by(email=data['email']).first():
        return jsonify({"msg": "Email already registered"}), 409
    user = User(
        email=data['email'],
        first_name=data.get('first_name', ''),
        last_name=data.get('last_name', ''),
        role='student'
    )
    user.set_password(data['password'])
    db.session.add(user)
    db.session.commit()
    log_audit('user_registered', user.id, f"Email: {user.email}")
    return jsonify(user.to_dict()), 201

@auth_bp.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    email = data.get('email')
    password = data.get('password')
    user = User.query.filter_by(email=email).first()
    if not user or not user.check_password(password):
        return jsonify({"msg": "Invalid email or password"}), 401
    if not user.is_verified:
        return jsonify({"msg": "Email not verified"}), 403
    access_token = create_access_token(identity=user.id, additional_claims={'role': user.role})
    refresh_token = create_refresh_token(identity=user.id)
    log_audit('user_login', user.id)
    return jsonify(access_token=access_token, refresh_token=refresh_token, user=user.to_dict()), 200

@auth_bp.route('/verify-email/<token>', methods=['GET'])
def verify_email(token):
    serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])
    try:
        email = serializer.loads(token, salt=app.config['TOKEN_SALT'], max_age=86400)
    except:
        return jsonify({"msg": "Invalid or expired token"}), 400
    user = User.query.filter_by(email=email).first()
    if not user:
        return jsonify({"msg": "User not found"}), 404
    user.is_verified = True
    db.session.commit()
    return jsonify({"msg": "Email verified successfully"}), 200

@auth_bp.route('/refresh', methods=['POST'])
@jwt_required(refresh=True)
def refresh():
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)
    if not user:
        return jsonify({"msg": "User not found"}), 404
    new_token = create_access_token(identity=user.id, additional_claims={'role': user.role})
    return jsonify(access_token=new_token), 200

@auth_bp.route('/forgot-password', methods=['POST'])
def forgot_password():
    data = request.get_json()
    email = data.get('email')
    user = User.query.filter_by(email=email).first()
    if user:
        serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])
        token = serializer.dumps(email, salt='password-reset')
        if app.config.get('MAIL_USERNAME'):
            from flask_mail import Message, Mail
            mail = Mail()
            reset_url = f"{request.host_url}api/auth/reset-password/{token}"
            msg = Message("Password Reset", recipients=[email])
            msg.body = f"Reset link: {reset_url}"
            try:
                mail.send(msg)
            except:
                pass
        return jsonify({"msg": "If email exists, a reset link has been sent.", "reset_token": token}), 200
    return jsonify({"msg": "If email exists, a reset link has been sent."}), 200

@auth_bp.route('/reset-password/<token>', methods=['POST'])
def reset_password(token):
    data = request.get_json()
    new_password = data.get('password')
    serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])
    try:
        email = serializer.loads(token, salt='password-reset', max_age=3600)
    except:
        return jsonify({"msg": "Invalid or expired token"}), 400
    user = User.query.filter_by(email=email).first()
    if not user:
        return jsonify({"msg": "User not found"}), 404
    user.set_password(new_password)
    db.session.commit()
    return jsonify({"msg": "Password reset successful"}), 200

# ---------- Admin Blueprint (Super Admin) ----------
admin_bp = Blueprint('admin', __name__)

@admin_bp.route('/users', methods=['GET'])
@role_required('super_admin', 'institution_admin')
def list_users():
    users = User.query.all()
    return jsonify([u.to_dict() for u in users]), 200

@admin_bp.route('/users/<int:user_id>', methods=['GET'])
@role_required('super_admin', 'institution_admin')
def get_user(user_id):
    user = User.query.get_or_404(user_id)
    return jsonify(user.to_dict()), 200

@admin_bp.route('/users/<int:user_id>', methods=['PUT'])
@role_required('super_admin')
def update_user(user_id):
    user = User.query.get_or_404(user_id)
    data = request.get_json()
    if 'role' in data:
        user.role = data['role']
    if 'institution_id' in data:
        user.institution_id = data['institution_id']
    db.session.commit()
    log_audit('user_updated', g.current_user_id, f"Updated user {user_id}")
    return jsonify(user.to_dict()), 200

@admin_bp.route('/users/<int:user_id>', methods=['DELETE'])
@role_required('super_admin')
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    db.session.delete(user)
    db.session.commit()
    log_audit('user_deleted', g.current_user_id, f"Deleted user {user_id}")
    return jsonify({"msg": "User deleted"}), 200

@admin_bp.route('/institutions', methods=['POST'])
@role_required('super_admin')
def create_institution():
    data = request.get_json()
    inst = Institution(
        name=data['name'], code=data.get('code'), type=data.get('type'),
        address=data.get('address'), logo_url=data.get('logo_url')
    )
    db.session.add(inst)
    db.session.commit()
    return jsonify({'id': inst.id, 'name': inst.name}), 201

@admin_bp.route('/institutions', methods=['GET'])
@role_required('super_admin')
def list_institutions():
    insts = Institution.query.all()
    return jsonify([{'id': i.id, 'name': i.name, 'code': i.code} for i in insts]), 200

# ---------- Institution Blueprint (Academic Hierarchy CRUD) ----------
institution_bp = Blueprint('institution', __name__)

# ---- Faculty ----
@institution_bp.route('/faculties', methods=['GET'])
@role_required('super_admin', 'institution_admin', 'hod')
def get_faculties():
    faculties = inst_query(Faculty).all()
    return jsonify([{'id': f.id, 'name': f.name, 'code': f.code, 'institution_id': f.institution_id} for f in faculties]), 200

@institution_bp.route('/faculties', methods=['POST'])
@role_required('super_admin', 'institution_admin')
def create_faculty():
    data = request.get_json()
    inst_id = get_institution_id()
    if inst_id is None and 'institution_id' not in data:
        return jsonify({"msg": "institution_id required for super admin"}), 400
    if inst_id is not None:
        data['institution_id'] = inst_id
    faculty = Faculty(name=data['name'], code=data.get('code'), institution_id=data['institution_id'])
    db.session.add(faculty)
    db.session.commit()
    log_audit('faculty_created', g.current_user_id, f"{faculty.name}")
    return jsonify({'id': faculty.id, 'name': faculty.name}), 201

@institution_bp.route('/faculties/<int:id>', methods=['PUT'])
@role_required('super_admin', 'institution_admin')
def update_faculty(id):
    faculty = Faculty.query.get_or_404(id)
    data = request.get_json()
    faculty.name = data.get('name', faculty.name)
    faculty.code = data.get('code', faculty.code)
    db.session.commit()
    return jsonify({'id': faculty.id, 'name': faculty.name}), 200

@institution_bp.route('/faculties/<int:id>', methods=['DELETE'])
@role_required('super_admin', 'institution_admin')
def delete_faculty(id):
    faculty = Faculty.query.get_or_404(id)
    db.session.delete(faculty)
    db.session.commit()
    return jsonify({"msg": "Deleted"}), 200

# ---- Department ----
@institution_bp.route('/departments', methods=['GET'])
@role_required('super_admin', 'institution_admin', 'hod')
def get_departments():
    depts = inst_query(Department).all()
    return jsonify([{'id': d.id, 'name': d.name, 'code': d.code, 'faculty_id': d.faculty_id, 'institution_id': d.institution_id} for d in depts]), 200

@institution_bp.route('/departments', methods=['POST'])
@role_required('super_admin', 'institution_admin')
def create_department():
    data = request.get_json()
    inst_id = get_institution_id()
    if inst_id is None and 'institution_id' not in data:
        return jsonify({"msg": "institution_id required"}), 400
    if inst_id is not None:
        data['institution_id'] = inst_id
    dept = Department(name=data['name'], code=data.get('code'), faculty_id=data['faculty_id'], institution_id=data['institution_id'])
    db.session.add(dept)
    db.session.commit()
    return jsonify({'id': dept.id, 'name': dept.name}), 201

@institution_bp.route('/departments/<int:id>', methods=['PUT'])
@role_required('super_admin', 'institution_admin')
def update_department(id):
    dept = Department.query.get_or_404(id)
    data = request.get_json()
    dept.name = data.get('name', dept.name)
    dept.code = data.get('code', dept.code)
    dept.faculty_id = data.get('faculty_id', dept.faculty_id)
    db.session.commit()
    return jsonify({'id': dept.id, 'name': dept.name}), 200

@institution_bp.route('/departments/<int:id>', methods=['DELETE'])
@role_required('super_admin', 'institution_admin')
def delete_department(id):
    dept = Department.query.get_or_404(id)
    db.session.delete(dept)
    db.session.commit()
    return jsonify({"msg": "Deleted"}), 200

# ---- Programme ----
@institution_bp.route('/programmes', methods=['GET'])
@role_required('super_admin', 'institution_admin', 'hod')
def get_programmes():
    inst_id = get_institution_id()
    if inst_id:
        programmes = Programme.query.join(Department).filter(Department.institution_id == inst_id).all()
    else:
        programmes = Programme.query.all()
    return jsonify([{'id': p.id, 'name': p.name, 'code': p.code, 'department_id': p.department_id, 'duration_years': p.duration_years} for p in programmes]), 200

@institution_bp.route('/programmes', methods=['POST'])
@role_required('super_admin', 'institution_admin')
def create_programme():
    data = request.get_json()
    prog = Programme(name=data['name'], code=data.get('code'), department_id=data['department_id'], duration_years=data.get('duration_years', 4))
    db.session.add(prog)
    db.session.commit()
    return jsonify({'id': prog.id, 'name': prog.name}), 201

@institution_bp.route('/programmes/<int:id>', methods=['PUT'])
@role_required('super_admin', 'institution_admin')
def update_programme(id):
    prog = Programme.query.get_or_404(id)
    data = request.get_json()
    prog.name = data.get('name', prog.name)
    prog.code = data.get('code', prog.code)
    prog.department_id = data.get('department_id', prog.department_id)
    prog.duration_years = data.get('duration_years', prog.duration_years)
    db.session.commit()
    return jsonify({'id': prog.id, 'name': prog.name}), 200

@institution_bp.route('/programmes/<int:id>', methods=['DELETE'])
@role_required('super_admin', 'institution_admin')
def delete_programme(id):
    prog = Programme.query.get_or_404(id)
    db.session.delete(prog)
    db.session.commit()
    return jsonify({"msg": "Deleted"}), 200

# ---- Academic Session ----
@institution_bp.route('/sessions', methods=['GET'])
@role_required('super_admin', 'institution_admin', 'hod')
def get_sessions():
    sessions = inst_query(AcademicSession).all()
    return jsonify([{'id': s.id, 'name': s.name, 'start_date': str(s.start_date), 'end_date': str(s.end_date), 'is_current': s.is_current} for s in sessions]), 200

@institution_bp.route('/sessions', methods=['POST'])
@role_required('super_admin', 'institution_admin')
def create_session():
    data = request.get_json()
    inst_id = get_institution_id()
    if inst_id is None and 'institution_id' not in data:
        return jsonify({"msg": "institution_id required"}), 400
    if inst_id is not None:
        data['institution_id'] = inst_id
    start_date = datetime.datetime.strptime(data['start_date'], '%Y-%m-%d').date() if data.get('start_date') else None
    end_date = datetime.datetime.strptime(data['end_date'], '%Y-%m-%d').date() if data.get('end_date') else None
    session = AcademicSession(
        name=data['name'], institution_id=data['institution_id'],
        start_date=start_date, end_date=end_date,
        is_current=data.get('is_current', False)
    )
    db.session.add(session)
    db.session.commit()
    return jsonify({'id': session.id, 'name': session.name}), 201

@institution_bp.route('/sessions/<int:id>', methods=['PUT'])
@role_required('super_admin', 'institution_admin')
def update_session(id):
    session = AcademicSession.query.get_or_404(id)
    data = request.get_json()
    session.name = data.get('name', session.name)
    if 'start_date' in data:
        session.start_date = datetime.datetime.strptime(data['start_date'], '%Y-%m-%d').date()
    if 'end_date' in data:
        session.end_date = datetime.datetime.strptime(data['end_date'], '%Y-%m-%d').date()
    session.is_current = data.get('is_current', session.is_current)
    db.session.commit()
    return jsonify({'id': session.id, 'name': session.name}), 200

@institution_bp.route('/sessions/<int:id>', methods=['DELETE'])
@role_required('super_admin', 'institution_admin')
def delete_session(id):
    session = AcademicSession.query.get_or_404(id)
    db.session.delete(session)
    db.session.commit()
    return jsonify({"msg": "Deleted"}), 200

# ---- Semester ----
@institution_bp.route('/semesters', methods=['GET'])
@role_required('super_admin', 'institution_admin', 'hod')
def get_semesters():
    inst_id = get_institution_id()
    if inst_id:
        semesters = Semester.query.join(AcademicSession).filter(AcademicSession.institution_id == inst_id).all()
    else:
        semesters = Semester.query.all()
    return jsonify([{'id': s.id, 'name': s.name, 'session_id': s.session_id, 'start_date': str(s.start_date), 'end_date': str(s.end_date)} for s in semesters]), 200

@institution_bp.route('/semesters', methods=['POST'])
@role_required('super_admin', 'institution_admin')
def create_semester():
    data = request.get_json()
    start_date = datetime.datetime.strptime(data['start_date'], '%Y-%m-%d').date() if data.get('start_date') else None
    end_date = datetime.datetime.strptime(data['end_date'], '%Y-%m-%d').date() if data.get('end_date') else None
    sem = Semester(name=data['name'], session_id=data['session_id'], start_date=start_date, end_date=end_date)
    db.session.add(sem)
    db.session.commit()
    return jsonify({'id': sem.id, 'name': sem.name}), 201

@institution_bp.route('/semesters/<int:id>', methods=['PUT'])
@role_required('super_admin', 'institution_admin')
def update_semester(id):
    sem = Semester.query.get_or_404(id)
    data = request.get_json()
    sem.name = data.get('name', sem.name)
    sem.session_id = data.get('session_id', sem.session_id)
    if 'start_date' in data:
        sem.start_date = datetime.datetime.strptime(data['start_date'], '%Y-%m-%d').date()
    if 'end_date' in data:
        sem.end_date = datetime.datetime.strptime(data['end_date'], '%Y-%m-%d').date()
    db.session.commit()
    return jsonify({'id': sem.id, 'name': sem.name}), 200

@institution_bp.route('/semesters/<int:id>', methods=['DELETE'])
@role_required('super_admin', 'institution_admin')
def delete_semester(id):
    sem = Semester.query.get_or_404(id)
    db.session.delete(sem)
    db.session.commit()
    return jsonify({"msg": "Deleted"}), 200

# ---- Course ----
@institution_bp.route('/courses', methods=['GET'])
@role_required('super_admin', 'institution_admin', 'hod', 'lecturer')
def get_courses():
    inst_id = get_institution_id()
    if inst_id:
        courses = Course.query.join(Department).filter(Department.institution_id == inst_id).all()
    else:
        courses = Course.query.all()
    return jsonify([{
        'id': c.id, 'code': c.code, 'title': c.title, 'credit_units': c.credit_units,
        'level': c.level, 'department_id': c.department_id, 'programme_id': c.programme_id,
        'semester_id': c.semester_id, 'lecturer_id': c.lecturer_id, 'students_count': c.students_count
    } for c in courses]), 200

@institution_bp.route('/courses', methods=['POST'])
@role_required('super_admin', 'institution_admin', 'hod')
def create_course():
    data = request.get_json()
    course = Course(
        code=data['code'], title=data['title'], credit_units=data.get('credit_units', 3),
        level=data.get('level'), department_id=data['department_id'],
        programme_id=data['programme_id'], semester_id=data['semester_id'],
        lecturer_id=data.get('lecturer_id'), students_count=data.get('students_count', 30)
    )
    db.session.add(course)
    db.session.commit()
    return jsonify({'id': course.id, 'code': course.code}), 201

@institution_bp.route('/courses/<int:id>', methods=['PUT'])
@role_required('super_admin', 'institution_admin', 'hod')
def update_course(id):
    course = Course.query.get_or_404(id)
    data = request.get_json()
    for field in ['code', 'title', 'credit_units', 'level', 'department_id', 'programme_id', 'semester_id', 'lecturer_id', 'students_count']:
        if field in data:
            setattr(course, field, data[field])
    db.session.commit()
    return jsonify({'id': course.id, 'code': course.code}), 200

@institution_bp.route('/courses/<int:id>', methods=['DELETE'])
@role_required('super_admin', 'institution_admin')
def delete_course(id):
    course = Course.query.get_or_404(id)
    db.session.delete(course)
    db.session.commit()
    return jsonify({"msg": "Deleted"}), 200

# ---- Lecturer ----
@institution_bp.route('/lecturers', methods=['GET'])
@role_required('super_admin', 'institution_admin', 'hod')
def get_lecturers():
    inst_id = get_institution_id()
    if inst_id:
        lecturers = Lecturer.query.join(Department).filter(Department.institution_id == inst_id).all()
    else:
        lecturers = Lecturer.query.all()
    result = []
    for l in lecturers:
        user = l.user
        result.append({
            'id': l.id, 'staff_id': l.staff_id, 'user_id': l.user_id,
            'name': f"{user.first_name} {user.last_name}" if user else "",
            'department_id': l.department_id, 'rank': l.rank,
            'max_daily_hours': l.max_daily_hours, 'max_weekly_hours': l.max_weekly_hours,
            'availability': l.availability_json
        })
    return jsonify(result), 200

@institution_bp.route('/lecturers', methods=['POST'])
@role_required('super_admin', 'institution_admin')
def create_lecturer():
    data = request.get_json()
    user = User.query.get(data['user_id'])
    if not user:
        return jsonify({"msg": "User not found"}), 404
    user.role = 'lecturer'
    lecturer = Lecturer(
        staff_id=data['staff_id'], user_id=data['user_id'],
        department_id=data['department_id'], rank=data.get('rank'),
        max_daily_hours=data.get('max_daily_hours', 6),
        max_weekly_hours=data.get('max_weekly_hours', 24),
        availability_json=json.dumps(data.get('availability', {}))
    )
    db.session.add(lecturer)
    db.session.commit()
    return jsonify({'id': lecturer.id, 'staff_id': lecturer.staff_id}), 201

@institution_bp.route('/lecturers/<int:id>', methods=['PUT'])
@role_required('super_admin', 'institution_admin')
def update_lecturer(id):
    lecturer = Lecturer.query.get_or_404(id)
    data = request.get_json()
    for field in ['staff_id', 'department_id', 'rank', 'max_daily_hours', 'max_weekly_hours']:
        if field in data:
            setattr(lecturer, field, data[field])
    if 'availability' in data:
        lecturer.availability_json = json.dumps(data['availability'])
    db.session.commit()
    return jsonify({'id': lecturer.id}), 200

@institution_bp.route('/lecturers/<int:id>/availability', methods=['PUT'])
@role_required('super_admin', 'institution_admin', 'lecturer')
def update_lecturer_availability(id):
    lecturer = Lecturer.query.get_or_404(id)
    data = request.get_json()
    lecturer.availability_json = json.dumps(data)
    db.session.commit()
    return jsonify({"msg": "Availability updated"}), 200

@institution_bp.route('/lecturers/<int:id>', methods=['DELETE'])
@role_required('super_admin', 'institution_admin')
def delete_lecturer(id):
    lecturer = Lecturer.query.get_or_404(id)
    db.session.delete(lecturer)
    db.session.commit()
    return jsonify({"msg": "Deleted"}), 200

# ---- Venue ----
@institution_bp.route('/venues', methods=['GET'])
@role_required('super_admin', 'institution_admin', 'hod')
def get_venues():
    venues = inst_query(Venue).all()
    return jsonify([{'id': v.id, 'code': v.code, 'name': v.name, 'capacity': v.capacity, 'building': v.building, 'resources': v.resources} for v in venues]), 200

@institution_bp.route('/venues', methods=['POST'])
@role_required('super_admin', 'institution_admin')
def create_venue():
    data = request.get_json()
    inst_id = get_institution_id()
    if inst_id is None and 'institution_id' not in data:
        return jsonify({"msg": "institution_id required"}), 400
    if inst_id is not None:
        data['institution_id'] = inst_id
    venue = Venue(
        code=data['code'], name=data.get('name'), capacity=data['capacity'],
        building=data.get('building'), resources=data.get('resources'),
        institution_id=data['institution_id']
    )
    db.session.add(venue)
    db.session.commit()
    return jsonify({'id': venue.id, 'code': venue.code}), 201

@institution_bp.route('/venues/<int:id>', methods=['PUT'])
@role_required('super_admin', 'institution_admin')
def update_venue(id):
    venue = Venue.query.get_or_404(id)
    data = request.get_json()
    for field in ['code', 'name', 'capacity', 'building', 'resources']:
        if field in data:
            setattr(venue, field, data[field])
    db.session.commit()
    return jsonify({'id': venue.id}), 200

@institution_bp.route('/venues/<int:id>', methods=['DELETE'])
@role_required('super_admin', 'institution_admin')
def delete_venue(id):
    venue = Venue.query.get_or_404(id)
    db.session.delete(venue)
    db.session.commit()
    return jsonify({"msg": "Deleted"}), 200

# ---------- Timetable Blueprint (AI Engine, Export, Approval) ----------
timetable_bp = Blueprint('timetable', __name__)

def get_time_slots_for_institution(institution_id=None):
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    periods = list(range(1, 9))   # 8 periods a day
    slots = [(d, p) for d in days for p in periods]
    return slots, days, periods

def generate_timetable(session_id, semester_id, user_id):
    courses = Course.query.filter_by(semester_id=semester_id).all()
    if not courses:
        return None, "No courses found for this semester"
    session = AcademicSession.query.get(session_id)
    if not session:
        return None, "Session not found"
    institution_id = session.institution_id
    venues = Venue.query.filter_by(institution_id=institution_id).all()
    if not venues:
        return None, "No venues available"
    lecturers = Lecturer.query.join(Department).filter(Department.institution_id == institution_id).all()
    lecturer_dict = {l.id: l for l in lecturers}
    for c in courses:
        if c.lecturer_id is None or c.lecturer_id not in lecturer_dict:
            dept_lecturers = [l for l in lecturers if l.department_id == c.department_id]
            c.lecturer_id = dept_lecturers[0].id if dept_lecturers else None
    valid_courses = [c for c in courses if c.lecturer_id is not None]
    if not valid_courses:
        return None, "No courses with assigned lecturers"

    all_slots, days, periods = get_time_slots_for_institution(institution_id)
    num_slots = len(all_slots)

    student_groups = {}
    for c in valid_courses:
        group_key = (c.programme_id, c.level)
        student_groups.setdefault(group_key, []).append(c)

    lecturer_courses = {}
    for c in valid_courses:
        lecturer_courses.setdefault(c.lecturer_id, []).append(c)

    model = cp_model.CpModel()
    course_vars = {}
    for c in valid_courses:
        slot_var = model.NewIntVar(0, num_slots - 1, f'slot_{c.id}')
        venue_var = model.NewIntVar(0, len(venues) - 1, f'venue_{c.id}')
        course_vars[c.id] = (slot_var, venue_var)

    # Lecturer clash
    for lecturer_id, courses_list in lecturer_courses.items():
        slot_vars = [course_vars[c.id][0] for c in courses_list]
        model.AddAllDifferent(slot_vars)

    # Student group clash
    for group, courses_list in student_groups.items():
        if len(courses_list) > 1:
            slot_vars = [course_vars[c.id][0] for c in courses_list]
            model.AddAllDifferent(slot_vars)

    # Venue clash
    for i in range(len(valid_courses)):
        ci = valid_courses[i]
        si_i, vi_i = course_vars[ci.id]
        for j in range(i + 1, len(valid_courses)):
            cj = valid_courses[j]
            si_j, vi_j = course_vars[cj.id]
            same_venue = model.NewBoolVar(f'same_venue_{ci.id}_{cj.id}')
            model.Add(vi_i == vi_j).OnlyEnforceIf(same_venue)
            model.Add(vi_i != vi_j).OnlyEnforceIf(same_venue.Not())
            model.Add(si_i != si_j).OnlyEnforceIf(same_venue)

    # Capacity
    for c in valid_courses:
        vi = course_vars[c.id][1]
        for v_idx, venue in enumerate(venues):
            if venue.capacity < c.students_count:
                model.Add(vi != v_idx)

    # Daily max hours per lecturer
    for lecturer_id, courses_list in lecturer_courses.items():
        l = lecturer_dict.get(lecturer_id)
        if not l:
            continue
        max_daily = l.max_daily_hours if l.max_daily_hours else 8
        for day_index, day_name in enumerate(days):
            day_start = day_index * len(periods)
            day_end = (day_index + 1) * len(periods) - 1
            on_day_vars = []
            for c in courses_list:
                slot_var = course_vars[c.id][0]
                is_on_day = model.NewBoolVar(f'on_day_{c.id}_{day_name}')
                model.Add(slot_var >= day_start).OnlyEnforceIf(is_on_day)
                model.Add(slot_var <= day_end).OnlyEnforceIf(is_on_day)
                on_day_vars.append(is_on_day)
            model.Add(sum(on_day_vars) <= max_daily)

    # Weekly max hours
    for lecturer_id, courses_list in lecturer_courses.items():
        l = lecturer_dict.get(lecturer_id)
        if not l:
            continue
        max_weekly = l.max_weekly_hours if l.max_weekly_hours else 40
        model.Add(len(courses_list) <= max_weekly)

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 50
    solver.parameters.num_search_workers = 8
    status = solver.Solve(model)

    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        return None, f"No feasible timetable found. Status: {solver.StatusName(status)}"

    assignments = []
    for c in valid_courses:
        slot_idx = solver.Value(course_vars[c.id][0])
        venue_idx = solver.Value(course_vars[c.id][1])
        day, period = all_slots[slot_idx]
        venue = venues[venue_idx]
        assignments.append({
            'course': c,
            'day': day,
            'period': period,
            'venue': venue,
            'lecturer_id': c.lecturer_id,
            'session_id': session_id,
            'semester_id': semester_id,
            'department_id': c.department_id
        })

    quality = 100.0   # base score (to be enhanced later)
    gen = TimetableGeneration(
        session_id=session_id, semester_id=semester_id,
        status='draft', created_by=user_id, quality_score=quality
    )
    db.session.add(gen)
    db.session.flush()

    for a in assignments:
        period_num = a['period']
        start_hour = 7 + period_num
        start_time = datetime.time(start_hour, 0)
        end_time = datetime.time(start_hour + 1, 0)
        slot = TimetableSlot(
            generation_id=gen.id,
            day=a['day'],
            period=period_num,
            start_time=start_time,
            end_time=end_time,
            course_id=a['course'].id,
            lecturer_id=a['lecturer_id'],
            venue_id=a['venue'].id,
            session_id=session_id,
            semester_id=semester_id,
            department_id=a['department_id']
        )
        db.session.add(slot)
    db.session.commit()
    return gen.id, None

@timetable_bp.route('/generate', methods=['POST'])
@role_required('institution_admin', 'hod')
def trigger_generate():
    data = request.get_json()
    session_id = data['session_id']
    semester_id = data['semester_id']
    gen_id, error = generate_timetable(session_id, semester_id, g.current_user_id)
    if error:
        return jsonify({"msg": error}), 400
    gen = TimetableGeneration.query.get(gen_id)
    return jsonify({
        "generation_id": gen.id,
        "quality_score": gen.quality_score,
        "status": gen.status
    }), 201

@timetable_bp.route('/generations', methods=['GET'])
@role_required('institution_admin', 'hod', 'lecturer', 'student')
def list_generations():
    gen_list = TimetableGeneration.query.order_by(TimetableGeneration.created_at.desc()).all()
    return jsonify([{
        'id': g.id, 'session_id': g.session_id, 'semester_id': g.semester_id,
        'status': g.status, 'quality_score': g.quality_score, 'created_at': str(g.created_at)
    } for g in gen_list]), 200

@timetable_bp.route('/slots', methods=['GET'])
@role_required('institution_admin', 'hod', 'lecturer', 'student')
def get_timetable_slots():
    generation_id = request.args.get('generation_id')
    if not generation_id:
        return jsonify({"msg": "generation_id required"}), 400
    slots = TimetableSlot.query.filter_by(generation_id=int(generation_id)).all()
    result = []
    for s in slots:
        course = Course.query.get(s.course_id)
        venue = Venue.query.get(s.venue_id)
        lecturer = Lecturer.query.get(s.lecturer_id)
        lecturer_user = User.query.get(lecturer.user_id) if lecturer else None
        result.append({
            'id': s.id,
            'day': s.day,
            'period': s.period,
            'start_time': str(s.start_time),
            'end_time': str(s.end_time),
            'course': {'code': course.code, 'title': course.title} if course else None,
            'lecturer': {'name': f"{lecturer_user.first_name} {lecturer_user.last_name}"} if lecturer_user else None,
            'venue': {'code': venue.code, 'name': venue.name} if venue else None,
            'department_id': s.department_id
        })
    return jsonify(result), 200

@timetable_bp.route('/slots/<int:slot_id>', methods=['PUT'])
@role_required('institution_admin')
def adjust_slot(slot_id):
    slot = TimetableSlot.query.get_or_404(slot_id)
    data = request.get_json()
    if 'day' in data:
        slot.day = data['day']
    if 'period' in data:
        slot.period = data['period']
    if 'venue_id' in data:
        slot.venue_id = data['venue_id']
    db.session.commit()
    log_audit('slot_adjusted', g.current_user_id, f"Slot {slot.id} modified")
    return jsonify({"msg": "Updated"}), 200

@timetable_bp.route('/approve/<int:gen_id>', methods=['PUT'])
@role_required('institution_admin', 'hod')
def approve_timetable(gen_id):
    gen = TimetableGeneration.query.get_or_404(gen_id)
    gen.status = 'approved'
    db.session.commit()
    log_audit('timetable_approved', g.current_user_id, f"Generation {gen_id} approved")
    return jsonify({"msg": "Timetable approved"}), 200

@timetable_bp.route('/conflicts/<int:gen_id>', methods=['GET'])
@role_required('institution_admin', 'hod')
def get_conflicts(gen_id):
    slots = TimetableSlot.query.filter_by(generation_id=gen_id).all()
    conflicts = []
    lect_map = {}
    for s in slots:
        key = (s.day, s.period, s.lecturer_id)
        if key in lect_map:
            conflicts.append({
                'type': 'lecturer_clash',
                'slot1_id': lect_map[key].id,
                'slot2_id': s.id,
                'detail': f"Lecturer double-booked on {s.day} period {s.period}"
            })
        else:
            lect_map[key] = s
    venue_map = {}
    for s in slots:
        key = (s.day, s.period, s.venue_id)
        if key in venue_map:
            conflicts.append({
                'type': 'venue_clash',
                'slot1_id': venue_map[key].id,
                'slot2_id': s.id,
                'detail': f"Venue double-booked on {s.day} period {s.period}"
            })
        else:
            venue_map[key] = s
    course_groups = {}
    for s in slots:
        course = Course.query.get(s.course_id)
        if not course:
            continue
        group = (course.programme_id, course.level)
        key = (s.day, s.period, group)
        if key in course_groups:
            conflicts.append({
                'type': 'student_clash',
                'slot1_id': course_groups[key].id,
                'slot2_id': s.id,
                'detail': f"Student group clash on {s.day} period {s.period}"
            })
        else:
            course_groups[key] = s
    return jsonify(conflicts), 200

@timetable_bp.route('/quality/<int:gen_id>', methods=['GET'])
@role_required('institution_admin', 'hod')
def get_quality(gen_id):
    gen = TimetableGeneration.query.get_or_404(gen_id)
    return jsonify({"quality_score": gen.quality_score}), 200

# ---- Export Endpoints ----
@timetable_bp.route('/export/pdf/<int:gen_id>', methods=['GET'])
@role_required('institution_admin', 'hod', 'lecturer', 'student')
def export_pdf(gen_id):
    slots = TimetableSlot.query.filter_by(generation_id=gen_id).order_by(TimetableSlot.day, TimetableSlot.period).all()
    if not slots:
        return jsonify({"msg": "No slots"}), 404
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()
    title = f"Timetable - Generation {gen_id}"
    elements.append(Paragraph(title, styles['Title']))
    elements.append(Spacer(1, 12))
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    periods = sorted(set(s.period for s in slots))
    data = [['Period'] + days]
    for p in periods:
        row = [str(p)]
        for d in days:
            cell = ""
            for s in slots:
                if s.day == d and s.period == p:
                    course = Course.query.get(s.course_id)
                    venue = Venue.query.get(s.venue_id)
                    cell = f"{course.code if course else '?'}\n{venue.code if venue else '?'}"
                    break
            row.append(cell)
        data.append(row)
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f"timetable_{gen_id}.pdf", mimetype='application/pdf')

@timetable_bp.route('/export/excel/<int:gen_id>', methods=['GET'])
@role_required('institution_admin', 'hod', 'lecturer', 'student')
def export_excel(gen_id):
    slots = TimetableSlot.query.filter_by(generation_id=gen_id).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timetable"
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    periods = sorted(set(s.period for s in slots))
    ws.append(['Period'] + days)
    for p in periods:
        row = [p]
        for d in days:
            cell_val = ""
            for s in slots:
                if s.day == d and s.period == p:
                    course = Course.query.get(s.course_id)
                    venue = Venue.query.get(s.venue_id)
                    cell_val = f"{course.code if course else '?'} ({venue.code if venue else '?'})"
                    break
            row.append(cell_val)
        ws.append(row)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f"timetable_{gen_id}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@timetable_bp.route('/export/csv/<int:gen_id>', methods=['GET'])
@role_required('institution_admin', 'hod', 'lecturer', 'student')
def export_csv(gen_id):
    slots = TimetableSlot.query.filter_by(generation_id=gen_id).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Day', 'Period', 'Course Code', 'Course Title', 'Lecturer', 'Venue'])
    for s in slots:
        course = Course.query.get(s.course_id)
        lecturer = Lecturer.query.get(s.lecturer_id)
        lecturer_user = User.query.get(lecturer.user_id) if lecturer else None
        venue = Venue.query.get(s.venue_id)
        writer.writerow([
            s.day, s.period,
            course.code if course else '', course.title if course else '',
            f"{lecturer_user.first_name} {lecturer_user.last_name}" if lecturer_user else '',
            venue.code if venue else ''
        ])
    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode('utf-8')), as_attachment=True,
                     download_name=f"timetable_{gen_id}.csv", mimetype='text/csv')

@timetable_bp.route('/export/image/<int:gen_id>', methods=['GET'])
@role_required('institution_admin', 'hod', 'lecturer', 'student')
def export_image(gen_id):
    slots = TimetableSlot.query.filter_by(generation_id=gen_id).all()
    img_width = 1000
    img_height = 600
    img = Image.new('RGB', (img_width, img_height), color='white')
    draw = ImageDraw.Draw(img)
    try:
        font = ImageFont.truetype("arial.ttf", 14)
    except:
        font = ImageFont.load_default()
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    periods = sorted(set(s.period for s in slots))
    col_width = img_width // (len(days)+1)
    row_height = 40
    draw.text((10, 10), "Timetable", fill='black', font=font)
    for i, day in enumerate(days):
        x = (i+1)*col_width
        draw.text((x, 40), day, fill='black', font=font)
    for j, p in enumerate(periods):
        y = 80 + j*row_height
        draw.text((10, y), str(p), fill='black', font=font)
        for i, d in enumerate(days):
            x = (i+1)*col_width
            text = ""
            for s in slots:
                if s.day == d and s.period == p:
                    course = Course.query.get(s.course_id)
                    text = course.code if course else "?"
                    break
            draw.text((x, y), text, fill='blue', font=font)
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    return send_file(buffer, mimetype='image/png', as_attachment=True, download_name=f"timetable_{gen_id}.png")

# ---------- Seeding and Run ----------
def seed_super_admin():
    if not User.query.filter_by(role='super_admin').first():
        admin = User(
            email='admin@timetable.com',
            first_name='Super',
            last_name='Admin',
            role='super_admin',
            is_verified=True
        )
        admin.set_password('Admin@123')
        db.session.add(admin)
        db.session.commit()
        log_audit('seed_super_admin', admin.id, "Default super admin created")

app = create_app()

if __name__ == '__main__':
    app.run(debug=True)
